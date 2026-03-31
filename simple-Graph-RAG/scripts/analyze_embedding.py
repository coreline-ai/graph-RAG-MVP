"""Analyze embedding quality of model_issue_dataset_10000.xlsx"""
import openpyxl
from io import BytesIO
import re
from collections import Counter, defaultdict
from datetime import datetime, date, timedelta

wb = openpyxl.load_workbook(
    'data/model_issue_dataset_10000.xlsx', data_only=True, read_only=True
)

print("=" * 70)
print("1. WORKBOOK 기본 정보")
print("=" * 70)
print(f"시트 수: {len(wb.sheetnames)}")
print(f"시트 이름: {wb.sheetnames}")

all_rows = []
all_headers = {}

for sheet in wb.worksheets:
    if sheet.sheet_state != "visible":
        continue
    values = list(sheet.iter_rows(values_only=True))
    if not values:
        continue
    headers = [str(h).strip() if h else "" for h in values[0]]
    all_headers[sheet.title] = headers
    print(f"\n--- Sheet: {sheet.title} ---")
    print(f"  헤더: {headers}")
    print(f"  데이터 행 수: {len(values) - 1}")

    for row in values[1:]:
        if not any(v not in (None, "") for v in row):
            continue
        row_dict = {}
        for i, h in enumerate(headers):
            if h and i < len(row):
                row_dict[h] = row[i]
        row_dict["_sheet"] = sheet.title
        all_rows.append(row_dict)

wb.close()

print(f"\n총 유효 데이터 행: {len(all_rows)}")

# Header aliases
HEADER_ALIASES = {
    "이슈": "모델 이슈 검토 사항",
    "진행": "진행(담당자)",
    "문제 원인 분석 결과": "문제점 분석 내용 (담당자 Comments)",
}

def canon(h):
    return HEADER_ALIASES.get(h, h)

canonical_rows = []
for r in all_rows:
    cr = {}
    for k, v in r.items():
        cr[canon(k)] = v
    canonical_rows.append(cr)

print("\n" + "=" * 70)
print("2. 데이터 필드 품질 분석")
print("=" * 70)

REQUIRED = [
    "모델 이슈 검토 사항", "등록일", "기본 확인내용", "기본 작업내용",
    "업무지시", "담당자", "업무시작일", "완료예정", "진행(담당자)", "완료일",
    "문제점 분석 내용 (담당자 Comments)"
]

for field in REQUIRED:
    filled = sum(1 for r in canonical_rows if r.get(field) not in (None, "", " "))
    pct = filled / len(canonical_rows) * 100 if canonical_rows else 0
    print(f"  {field:45s} : {filled:>6d}/{len(canonical_rows):>6d} ({pct:5.1f}%)")

print("\n" + "=" * 70)
print("3. 이슈 제목 분석 (임베딩 핵심 필드)")
print("=" * 70)

titles = [
    str(r.get("모델 이슈 검토 사항", "")).strip()
    for r in canonical_rows
    if r.get("모델 이슈 검토 사항")
]
title_lengths = [len(t) for t in titles]
if title_lengths:
    print(f"  총 이슈 수: {len(titles)}")
    print(f"  제목 길이 - 최소: {min(title_lengths)}, 최대: {max(title_lengths)}, "
          f"평균: {sum(title_lengths)/len(title_lengths):.1f}")

    short = [t for t in titles if len(t) < 10]
    print(f"  짧은 제목 (< 10자): {len(short)}건")
    for s in short[:5]:
        print(f"    - {repr(s)}")

    title_counts = Counter(titles)
    dupes = {t: c for t, c in title_counts.items() if c > 1}
    print(f"  중복 제목: {len(dupes)}건 (총 {sum(dupes.values())}행)")
    for t, c in sorted(dupes.items(), key=lambda x: -x[1])[:5]:
        print(f"    - [{c}회] {repr(t[:60])}")

print("\n" + "=" * 70)
print("4. 문제점 분석 내용 (analysis) 품질")
print("=" * 70)

analyses = [
    str(r.get("문제점 분석 내용 (담당자 Comments)", "")).strip()
    for r in canonical_rows
]
non_empty = [a for a in analyses if a and a != "None"]
analysis_lengths = [len(a) for a in non_empty]
if analysis_lengths:
    print(f"  분석 내용 있는 행: {len(non_empty)}/{len(canonical_rows)} "
          f"({len(non_empty)/len(canonical_rows)*100:.1f}%)")
    print(f"  분석 길이 - 최소: {min(analysis_lengths)}, 최대: {max(analysis_lengths)}, "
          f"평균: {sum(analysis_lengths)/len(analysis_lengths):.1f}")

    buckets = [(0, 100), (100, 300), (300, 600), (600, 1000), (1000, 3000), (3000, 999999)]
    for lo, hi in buckets:
        count = sum(1 for ln in analysis_lengths if lo <= ln < hi)
        label = f"{lo}-{hi}" if hi != 999999 else f"{lo}+"
        bar = "#" * int(count / max(len(analysis_lengths), 1) * 50)
        print(f"    {label:>10s}: {count:>5d} {bar}")
else:
    print("  분석 내용 없음!")

print("\n" + "=" * 70)
print("5. 청크 텍스트 시뮬레이션 (임베딩 입력 분석)")
print("=" * 70)

EXCEL_ROW_MAX_CHARS = 600
overview_chunks = 0
analysis_chunks = 0
chunk_texts = []
overview_lengths = []
analysis_flow_lengths = []

for r in canonical_rows:
    title = str(r.get("모델 이슈 검토 사항", "")).strip()
    if not title or title == "None":
        continue

    reg_date = r.get("등록일", "")
    if isinstance(reg_date, datetime):
        reg_date = reg_date.date().isoformat()
    elif isinstance(reg_date, date):
        reg_date = reg_date.isoformat()
    else:
        reg_date = str(reg_date) if reg_date else ""

    check = str(r.get("기본 확인내용", "") or "").strip()
    work = str(r.get("기본 작업내용", "") or "").strip()
    instr = str(r.get("업무지시", "") or "").strip()
    assignee = str(r.get("담당자", "") or "unassigned").strip()
    status_raw = str(r.get("진행(담당자)", "") or "").strip()
    analysis_text = str(r.get("문제점 분석 내용 (담당자 Comments)", "") or "").strip()
    if analysis_text == "None":
        analysis_text = ""

    lines = [
        f"[이슈] {title}",
        f"[등록일] {reg_date}",
        f"[기본 확인내용] {check}",
        f"[기본 작업내용] {work}",
        f"[업무지시] {instr}",
        f"[담당자] {assignee}",
        f"[진행] {status_raw}",
    ]
    overview_text = "\n".join(lines)
    single_text = overview_text + (
        f"\n[문제점 분석 내용] {analysis_text}" if analysis_text else ""
    )

    should_split = analysis_text and len(single_text) > EXCEL_ROW_MAX_CHARS

    if should_split:
        overview_chunks += 1
        overview_lengths.append(len(overview_text))
        chunk_texts.append(overview_text)
        analysis_chunks += 1
        analysis_flow_lengths.append(len(analysis_text))
        chunk_texts.append(analysis_text[:1000])
    else:
        overview_chunks += 1
        overview_lengths.append(len(single_text))
        chunk_texts.append(single_text)

total_chunks = overview_chunks + analysis_chunks
print(f"  총 청크 수 (예상): {total_chunks}")
print(f"    - Overview 청크: {overview_chunks}")
print(f"    - Analysis Flow 청크: {analysis_chunks}")

if overview_lengths:
    print(f"\n  Overview 청크 길이:")
    print(f"    최소: {min(overview_lengths)}, 최대: {max(overview_lengths)}, "
          f"평균: {sum(overview_lengths)/len(overview_lengths):.1f}")

if analysis_flow_lengths:
    print(f"  Analysis Flow 청크 길이:")
    print(f"    최소: {min(analysis_flow_lengths)}, 최대: {max(analysis_flow_lengths)}, "
          f"평균: {sum(analysis_flow_lengths)/len(analysis_flow_lengths):.1f}")

all_chunk_lengths = [len(t) for t in chunk_texts]
if all_chunk_lengths:
    print(f"\n  전체 청크 길이 분포:")
    buckets2 = [(0, 100), (100, 200), (200, 400), (400, 600), (600, 1000), (1000, 999999)]
    for lo, hi in buckets2:
        count = sum(1 for ln in all_chunk_lengths if lo <= ln < hi)
        label = f"{lo}-{hi}" if hi != 999999 else f"{lo}+"
        bar = "#" * int(count / max(len(all_chunk_lengths), 1) * 50)
        print(f"    {label:>10s}: {count:>5d} {bar}")

print("\n" + "=" * 70)
print("6. 담당자/상태 분포 (메타데이터 필터링 품질)")
print("=" * 70)

assignees = Counter(
    str(r.get("담당자", "")).strip()
    for r in canonical_rows
    if r.get("담당자")
)
print(f"  담당자 수: {len(assignees)}")
for a, c in assignees.most_common(10):
    print(f"    {a:20s}: {c:>5d}")

statuses = Counter(
    str(r.get("진행(담당자)", "")).strip()
    for r in canonical_rows
    if r.get("진행(담당자)")
)
print(f"\n  진행 상태 수: {len(statuses)}")
for s, c in statuses.most_common(15):
    print(f"    {s[:40]:40s}: {c:>5d}")

print("\n" + "=" * 70)
print("7. 임베딩 품질 위험 요소 종합 분석")
print("=" * 70)

issues_found = 0

very_short = sum(1 for ln in all_chunk_lengths if ln < 50)
if very_short:
    issues_found += 1
    print(f"  [CRITICAL] 매우 짧은 청크 (< 50자): {very_short}건")
    print(f"    -> 임베딩 벡터가 의미를 충분히 담지 못할 수 있음")
else:
    print(f"  [OK] 매우 짧은 청크 없음")

very_long = sum(1 for ln in all_chunk_lengths if ln > 2000)
if very_long:
    issues_found += 1
    print(f"  [WARNING] 매우 긴 청크 (> 2000자): {very_long}건")
    print(f"    -> BGE-M3 최대 토큰 8192이지만, 너무 긴 텍스트는 의미 희석 가능")
else:
    print(f"  [OK] 과도하게 긴 청크 없음")

chunk_dupes = Counter(chunk_texts)
exact_dupes = {t: c for t, c in chunk_dupes.items() if c > 1}
if exact_dupes:
    issues_found += 1
    total_dupe_rows = sum(exact_dupes.values())
    print(f"  [WARNING] 완전 중복 청크: {len(exact_dupes)}종류 (총 {total_dupe_rows}행)")
    print(f"    -> 동일 벡터가 여러 개 저장되어 검색 결과 다양성 저하")
else:
    print(f"  [OK] 완전 중복 청크 없음")

no_analysis = sum(1 for a in analyses if not a or a == "None")
ratio = no_analysis / len(analyses) * 100 if analyses else 0
if ratio > 50:
    issues_found += 1
    print(f"  [WARNING] 분석 내용 없는 행: {no_analysis}/{len(analyses)} ({ratio:.1f}%)")
    print(f"    -> 분석 없는 이슈는 overview만으로 검색되며, 상세 매칭 어려움")
else:
    print(f"  [INFO] 분석 내용 있는 비율: {100-ratio:.1f}%")

no_date = sum(1 for r in canonical_rows if not r.get("등록일"))
if no_date:
    print(f"  [WARNING] 등록일 누락: {no_date}건")
else:
    print(f"  [OK] 등록일 누락 없음")

unassigned = sum(
    1 for r in canonical_rows
    if str(r.get("담당자", "")).strip() in ("", "None", "unassigned")
)
if unassigned:
    print(f"  [INFO] 담당자 미지정: {unassigned}건")

korean_pattern = re.compile(r"[가-힣]")
non_korean = sum(1 for t in chunk_texts if not korean_pattern.search(t))
if non_korean:
    issues_found += 1
    print(f"  [CRITICAL] 한국어 미포함 청크: {non_korean}건")
    print(f"    -> BGE-M3은 다국어 모델이지만, 의미 없는 텍스트일 가능성")
else:
    print(f"  [OK] 모든 청크에 한국어 포함")

# Semantic diversity check - check unique first 50 chars
first50 = Counter(t[:50] for t in chunk_texts)
high_prefix_dupes = sum(1 for _, c in first50.items() if c > 3)
if high_prefix_dupes:
    print(f"  [WARNING] 접두사 중복 (첫 50자 동일) 패턴: {high_prefix_dupes}종류")
    print(f"    -> 유사한 청크가 많으면 벡터 공간에서 클러스터링 편향 발생")

print(f"\n  === 총 {issues_found}개 위험 요소 발견 ===")

print("\n" + "=" * 70)
print("8. BGE-M3 임베딩 모델 적합성 평가")
print("=" * 70)
print(f"  모델: BAAI/bge-m3")
print(f"  차원: 1024")
print(f"  최대 토큰: 8192")
print(f"  특성: 다국어(한국어 포함), Dense + Sparse + ColBERT 지원")
print(f"  normalize_embeddings: True (코사인 유사도 최적화)")
print()
avg_chars = sum(all_chunk_lengths) / len(all_chunk_lengths) if all_chunk_lengths else 0
est_tokens = avg_chars * 0.5  # Korean ~0.5 tokens per char for BGE-M3
print(f"  평균 청크 길이: {avg_chars:.0f}자 (추정 ~{est_tokens:.0f} 토큰)")
if est_tokens < 8192:
    print(f"  [OK] 토큰 한도 내 (8192)")
else:
    print(f"  [CRITICAL] 토큰 한도 초과!")

print(f"\n  벡터 저장소: PostgreSQL + pgvector (HNSW 인덱스, cosine)")
print(f"  인덱스: vector_cosine_ops -> normalize_embeddings=True와 일치 [OK]")

print("\n" + "=" * 70)
print("9. 샘플 청크 텍스트")
print("=" * 70)
samples = [0, len(chunk_texts)//4, len(chunk_texts)//2, len(chunk_texts)-1]
for idx in samples:
    if 0 <= idx < len(chunk_texts):
        print(f"\n--- 청크 #{idx} (길이: {len(chunk_texts[idx])}) ---")
        print(chunk_texts[idx][:400])
        if len(chunk_texts[idx]) > 400:
            print("... (truncated)")
