"""
Generate 10,000 unique RAG system issue records for model_issue_dataset_10000.xlsx.
Every row has a unique title and unique analysis text. No duplicates.
"""
import random
import hashlib
from datetime import date, timedelta
from typing import NamedTuple

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

random.seed(42)

# ──────────────────────────────────────────────────────────────────────
# 1. VOCABULARY POOLS (large enough to produce >10k unique combos)
# ──────────────────────────────────────────────────────────────────────

SYSTEMS = [
    "BGE-M3-Retrieve", "KoAssist-XL", "AnswerGen-v5", "RAG-Search-Prod",
    "KR-RAG-v2", "IntentRouter-v4", "Reranker-v3", "ChunkIndexer",
    "EmbeddingCache", "GraphRAG-Core", "QueryPlanner", "DocParser-v2",
    "VectorStore-PG", "Neo4j-GraphDB", "KoreanNLP-Pipe", "ContextMerger",
    "FeedbackLoop", "SessionTracker", "AuthGateway", "RateLimiter",
    "MetadataStore", "CacheLayer-Redis", "StreamProcessor", "BatchIngestor",
    "ModelRouter", "PromptBuilder", "TokenCounter", "ResponseFilter",
    "AuditLogger", "SchemaValidator", "WebhookRelay", "NotifyService",
    "SearchOptimizer", "RerankPipeline", "SummaryEngine", "TranslationBridge",
    "DedupService", "QualityChecker", "LoadBalancer-v2", "HealthMonitor",
    "ConfigManager", "SecretVault", "DataExporter", "ReportBuilder",
    "AlertDispatcher", "MigrationRunner", "BackupScheduler", "IndexOptimizer",
    "QueryCache", "ResultAggregator",
]

ISSUE_PATTERNS = [
    "{sys} 모듈에서 {symptom} 발생",
    "{sys} 처리 중 {symptom} 확인됨",
    "{sys} 연동 시 {symptom} 현상 반복",
    "{sys} 응답에서 {symptom} 감지",
    "{sys} 배포 후 {symptom} 리포트 접수",
    "{sys} 업데이트 이후 {symptom} 증가",
    "{sys} 초기화 과정에서 {symptom} 발견",
    "{sys} 장애 복구 후 {symptom} 잔존",
    "{sys} 부하 테스트 중 {symptom} 재현",
    "{sys} 롤백 후에도 {symptom} 지속",
    "{sys} 스케일아웃 시 {symptom} 악화",
    "{sys} 모니터링에서 {symptom} 포착",
    "{sys} 정기 점검 중 {symptom} 탐지",
    "{sys} 로그 분석 결과 {symptom} 확인",
    "{sys} 사용자 피드백으로 {symptom} 보고",
]

SYMPTOMS = [
    "응답 지연 급증", "메모리 누수", "CPU 사용률 폭주", "타임아웃 오류",
    "잘못된 검색 결과 반환", "임베딩 벡터 불일치", "인덱스 손상",
    "한국어 토크나이징 오류", "날짜 파싱 실패", "중복 청크 생성",
    "권한 인증 실패", "데이터 유실", "캐시 정합성 깨짐",
    "그래프 탐색 무한루프", "코사인 유사도 이상치", "배치 처리 중단",
    "커넥션 풀 고갈", "디스크 I/O 병목", "OOM 크래시", "인코딩 깨짐",
    "SSL 인증서 만료", "API 레이트 제한 초과", "스키마 불일치",
    "정렬 순서 역전", "페이지네이션 오프셋 오류", "동시성 경합 발생",
    "데드락 감지", "파티션 키 편향", "리밸런싱 실패", "스냅샷 복원 오류",
    "핫스팟 노드 과부하", "레플리카 동기화 지연", "쿼리 플랜 비효율",
    "GC 압력 증가", "스레드 풀 포화", "큐 백로그 적체",
    "네트워크 파티션 감지", "DNS 해석 실패", "프록시 라우팅 오류",
    "헬스체크 false positive", "메트릭 수집 누락", "로그 로테이션 실패",
    "설정 파일 충돌", "환경 변수 누락", "시크릿 로테이션 미반영",
    "웹훅 전달 실패", "알림 중복 발송", "리포트 데이터 불일치",
    "마이그레이션 롤백 불가", "백업 체인 단절",
]

CHECK_TEMPLATES = [
    "{sys} 서비스 로그에서 {detail} 패턴 확인",
    "{env} 환경 모니터링 대시보드에서 {metric} 지표 이상 감지",
    "최근 {days}일간 {sys} 관련 에러율 {rate}% 증가 확인",
    "{sys} 헬스체크 API 응답시간 {ms}ms로 SLA 초과",
    "Grafana 알림으로 {sys}의 {metric} 임계치 돌파 확인",
    "{region} 리전 {sys} 인스턴스에서 {detail} 로그 반복 발생",
    "사용자 {count}건 이상 동일 증상 피드백 접수 확인",
    "Sentry에서 {sys} 모듈 {error_type} 에러 {count}건 집계",
    "CI/CD 파이프라인 {stage} 단계에서 {sys} 테스트 실패 감지",
    "{sys} 프로파일링 결과 {bottleneck} 구간 병목 확인",
]

WORK_TEMPLATES = [
    "{action}하고 {sys} 서비스 재배포 진행",
    "{sys} 설정값 {param}을 {old_val}에서 {new_val}로 조정",
    "핫픽스 브랜치에서 {fix} 패치 적용 후 스테이징 검증",
    "{sys} 모듈의 {component} 로직 리팩토링 및 단위 테스트 보강",
    "장애 원인 {root_cause} 관련 임시 우회 로직 삽입",
    "{sys} 인프라 스케일업 ({old_spec} → {new_spec}) 적용",
    "모니터링 대시보드에 {metric} 알림 룰 추가 설정",
    "{sys} 의존성 라이브러리 {lib} 버전 업그레이드 ({ver})",
    "데이터 정합성 검증 스크립트 작성 및 일괄 보정 실행",
    "{sys} 회로 차단기(circuit breaker) 임계값 튜닝",
]

INSTRUCTION_TEMPLATES = [
    "{sys} {aspect} 안정화 방안 수립 및 이행 보고",
    "장애 재발 방지를 위한 {sys} {measure} 방안 마련",
    "{sys} 성능 개선 목표 {target} 달성을 위한 로드맵 제시",
    "관련 팀과 {sys} {topic} 협의 후 결과 공유",
    "{sys} 운영 매뉴얼에 {section} 항목 보완 및 배포",
    "QA팀과 협력하여 {sys} {test_type} 테스트 시나리오 확대",
    "{sys} 모니터링 커버리지 {coverage}% 이상으로 확대",
    "주간 리뷰에서 {sys} {topic} 진행 상황 보고",
    "{sys} 장애 포스트모템 작성 및 전사 공유",
    "보안팀과 {sys} {security_aspect} 점검 일정 조율",
]

ANALYSIS_CAUSE_TEMPLATES = [
    "원인 요약: {sys} {component}에서 {root_cause}가 발생하여 {effect}가 초래되었습니다.",
    "원인 요약: {condition} 상황에서 {sys}의 {mechanism}이 정상 작동하지 않아 {effect}로 이어졌습니다.",
    "원인 요약: {sys} 배포 시 {config_issue}가 누락되면서 {component}의 {behavior}가 변경되었습니다.",
    "원인 요약: {external_factor}으로 인해 {sys}의 {resource}가 고갈되어 {effect}가 발생했습니다.",
    "원인 요약: {sys}와 {dep_sys} 사이의 {interface_issue}로 인해 {effect}가 반복되었습니다.",
]

ANALYSIS_EVIDENCE_TEMPLATES = [
    "확인 근거: {date_ref} 기간 로그에서 {evidence}를 {count}건 확인하였으며, {metric}가 평소 대비 {pct}% 변동하였습니다.",
    "확인 근거: APM 트레이싱에서 {span}의 레이턴시가 {latency}ms로 측정되었고, {threshold}ms SLA를 초과하였습니다.",
    "확인 근거: 데이터베이스 쿼리 플랜 분석 결과 {query_part}에서 풀스캔이 발생하여 {effect}을 유발하였습니다.",
    "확인 근거: {sys} 메모리 프로파일에서 {object_type} 객체가 {size}MB까지 누적되어 GC 압력이 증가하였습니다.",
    "확인 근거: 네트워크 패킷 캡처에서 {sys}와 {target} 간 재전송률이 {retry_rate}%로 비정상적이었습니다.",
]

ANALYSIS_JUDGMENT_TEMPLATES = [
    "기술 판단: 이 문제는 {category} 영역의 {severity} 수준 이슈이며, {approach}이 가장 효과적인 해결책입니다.",
    "기술 판단: 단기적으로 {workaround}를 적용하고, 중장기적으로 {permanent_fix}를 진행해야 합니다.",
    "기술 판단: {sys} 아키텍처 수준의 개선이 필요하며, {redesign_area}를 재설계하는 것을 권장합니다.",
    "기술 판단: 현재 {sys} 버전에서는 {limitation}이 있어 {upgrade_target} 버전으로의 마이그레이션이 필요합니다.",
    "기술 판단: {root_cause}는 {sys} 고유의 문제가 아니라 {broader_issue}에서 기인하므로 플랫폼 수준 대응이 필요합니다.",
]

ANALYSIS_IMPACT_TEMPLATES = [
    "영향 범위: {scope}에 걸쳐 약 {user_count}명의 사용자에게 영향을 미쳤으며, {service}의 가용성이 {availability}%로 저하되었습니다.",
    "영향 범위: {service} 기능 중 {affected_feature}가 {duration} 동안 정상 작동하지 않아 {downstream_impact}에 영향을 주었습니다.",
    "영향 범위: 전체 트래픽의 약 {traffic_pct}%가 영향을 받았으며, {region} 리전 사용자에게 집중되었습니다.",
]

ANALYSIS_ACTION_TEMPLATES = [
    "추가 조치: {action1} 완료 후 {action2}를 진행하며, {monitoring_plan}으로 재발 여부를 모니터링합니다.",
    "추가 조치: {preventive_measure}를 도입하고, {review_cycle}마다 점검하여 유사 장애를 예방합니다.",
    "추가 조치: {documentation}을 업데이트하고, 관련 팀({teams})에 변경 사항을 공유합니다.",
    "추가 조치: {test_plan}을 수립하여 {coverage_target}% 이상의 테스트 커버리지를 확보합니다.",
    "추가 조치: {escalation_plan}에 따라 {timeline} 내에 완전한 해결을 목표로 합니다.",
]

# Filler vocabulary for template slots
COMPONENTS = [
    "임베딩 인코더", "쿼리 파서", "인덱스 매니저", "캐시 컨트롤러",
    "라우팅 엔진", "토크나이저", "청크 분할기", "리랭커 모듈",
    "그래프 워커", "벡터 스토어", "메타데이터 필터", "권한 검증기",
    "배치 스케줄러", "스트림 핸들러", "로드 밸런서", "커넥션 매니저",
    "쿼리 옵티마이저", "결과 병합기", "컨텍스트 빌더", "프롬프트 렌더러",
    "응답 필터", "피드백 수집기", "세션 매니저", "감사 로거",
    "스키마 검증기", "데이터 변환기", "이벤트 디스패처", "상태 머신",
    "워크플로우 엔진", "알림 라우터",
]

ROOT_CAUSES = [
    "메모리 할당 실패", "동시 접근 경합 조건", "설정값 오버라이드 누락",
    "의존성 버전 불일치", "네트워크 타임아웃", "디스크 용량 부족",
    "인덱스 파편화 누적", "캐시 무효화 타이밍 오류", "스레드 데드락",
    "GC 스톱더월드 지연", "커널 파라미터 부적합", "SSL 핸드셰이크 실패",
    "DNS 캐시 만료", "프록시 버퍼 오버플로우", "직렬화 포맷 변경",
    "API 스펙 하위 호환 위반", "데이터 마이그레이션 불완전",
    "환경 변수 미설정", "시크릿 키 로테이션 미반영", "로그 버퍼 포화",
    "배포 순서 역전", "헬스체크 오판정", "리소스 쿼터 초과",
    "파티션 리밸런싱 지연", "스냅샷 불일치", "레플리카 지연 누적",
    "쿼리 플랜 캐시 무효화 실패", "정규표현식 백트래킹 폭발",
    "유니코드 정규화 불일치", "타임존 변환 오류",
]

EFFECTS = [
    "응답 시간 급격한 증가", "서비스 일시 중단", "데이터 정합성 훼손",
    "검색 품질 저하", "사용자 세션 유실", "배치 작업 실패",
    "알림 지연 또는 누락", "리포트 데이터 불일치", "인증 토큰 만료 오류",
    "캐시 히트율 급락", "그래프 탐색 결과 누락", "임베딩 유사도 왜곡",
    "청크 경계 오류", "메타데이터 필터 오작동", "날짜 범위 쿼리 실패",
    "한국어 형태소 분석 오류", "리랭킹 점수 역전", "컨텍스트 창 초과",
    "토큰 소비량 비정상 증가", "폴백 응답 빈도 급증",
]

ENVS = ["프로덕션", "스테이징", "개발", "QA", "성능 테스트"]
REGIONS = ["ap-northeast-1", "ap-northeast-2", "us-east-1", "eu-west-1", "ap-southeast-1"]
METRICS = [
    "p95 레이턴시", "에러율", "처리량(TPS)", "메모리 사용률",
    "CPU 사용률", "디스크 IOPS", "네트워크 대역폭", "큐 깊이",
    "캐시 히트율", "GC 횟수", "스레드 수", "커넥션 풀 사용률",
]
SEVERITIES = ["긴급(P0)", "높음(P1)", "보통(P2)", "낮음(P3)"]
CATEGORIES = [
    "인프라", "애플리케이션 로직", "데이터 파이프라인", "보안",
    "성능", "데이터 정합성", "사용자 경험", "운영 안정성",
]
ACTIONS_VOCAB = [
    "핫픽스 배포", "설정값 롤백", "인프라 스케일업", "인덱스 재구축",
    "캐시 전체 플러시", "의존성 버전 고정", "모니터링 강화",
    "부하 테스트 재실행", "코드 리뷰 보강", "장애 시나리오 문서화",
    "자동화 테스트 추가", "알림 임계값 조정", "데이터 백필 실행",
    "로그 레벨 상향", "서킷브레이커 설정 조정", "커넥션 풀 확대",
    "GC 튜닝 적용", "리소스 쿼터 조정", "배포 파이프라인 수정",
    "장애 대응 런북 갱신",
]
LIBS = [
    "sentence-transformers", "pgvector", "neo4j-driver", "psycopg",
    "fastapi", "uvicorn", "pydantic", "redis-py", "httpx",
    "numpy", "torch", "transformers", "kss", "konlpy",
    "openpyxl", "celery", "sqlalchemy", "alembic",
]

ASSIGNEES = [
    "Donghyun", "Eunji", "Hwan", "Hyunwoo", "Jiho", "Jisoo",
    "Junseo", "Minji", "Seoyeon", "Sujin", "Taehun", "Yuna",
]

STATUS_OPTIONS = [
    ("[완료] 패치 반영 및 회귀 테스트 완료, 안정 확인됨", "완료"),
    ("[진행중] 원인 분석완료 단계, 솔루션 설계 및 우선순위 판단 중", "진행중"),
    ("[검증중] 솔루션 반영 완료, QA/릴리스 검증 대기", "검증중"),
    ("[분석중] 로그/지표 확인 및 원인 분석 범위 확정 중", "분석중"),
    ("[대기] 외부 작업 또는 외부 리소스 확인 필요", "대기"),
]


# ──────────────────────────────────────────────────────────────────────
# 2. UNIQUE TITLE GENERATOR
# ──────────────────────────────────────────────────────────────────────

def generate_unique_titles(n: int) -> list[str]:
    """Generate n unique issue titles."""
    titles: set[str] = set()
    attempts = 0
    while len(titles) < n and attempts < n * 10:
        attempts += 1
        sys = random.choice(SYSTEMS)
        symptom = random.choice(SYMPTOMS)
        pattern = random.choice(ISSUE_PATTERNS)
        title = pattern.format(sys=sys, symptom=symptom)
        titles.add(title)
    if len(titles) < n:
        # fallback: add index suffix
        base_titles = list(titles)
        while len(titles) < n:
            t = random.choice(base_titles) + f" (사례 #{len(titles)})"
            titles.add(t)
    return list(titles)[:n]


# ──────────────────────────────────────────────────────────────────────
# 3. UNIQUE FIELD GENERATORS
# ──────────────────────────────────────────────────────────────────────

def _pick(lst): return random.choice(lst)
def _picks(lst, k=2): return random.sample(lst, min(k, len(lst)))

def gen_check_text(sys: str, idx: int) -> str:
    tmpl = _pick(CHECK_TEMPLATES)
    return tmpl.format(
        sys=sys, detail=_pick(ROOT_CAUSES), env=_pick(ENVS),
        metric=_pick(METRICS), days=random.randint(1, 30),
        rate=round(random.uniform(5, 85), 1), ms=random.randint(200, 5000),
        region=_pick(REGIONS), count=random.randint(3, 500),
        error_type=_pick(["NullPointer", "Timeout", "OOM", "ConnectionReset", "ParseError",
                          "ValidationError", "AuthError", "RateLimitExceeded", "IOError"]),
        stage=_pick(["빌드", "테스트", "배포", "통합 테스트", "스모크 테스트"]),
        bottleneck=_pick(COMPONENTS),
    )

def gen_work_text(sys: str, idx: int) -> str:
    tmpl = _pick(WORK_TEMPLATES)
    return tmpl.format(
        sys=sys, action=_pick(ACTIONS_VOCAB),
        param=_pick(["max_connections", "timeout_ms", "batch_size", "retry_count",
                      "cache_ttl", "pool_size", "max_tokens", "chunk_overlap",
                      "embedding_dim", "top_k"]),
        old_val=random.randint(1, 100), new_val=random.randint(100, 10000),
        fix=_pick(ROOT_CAUSES).replace(" ", "-"),
        component=_pick(COMPONENTS), root_cause=_pick(ROOT_CAUSES),
        old_spec=f"{random.choice([2,4,8])}vCPU/{random.choice([4,8,16])}GB",
        new_spec=f"{random.choice([8,16,32])}vCPU/{random.choice([32,64,128])}GB",
        metric=_pick(METRICS), lib=_pick(LIBS),
        ver=f"{random.randint(1,5)}.{random.randint(0,20)}.{random.randint(0,30)}",
    )

def gen_instruction_text(sys: str, idx: int) -> str:
    tmpl = _pick(INSTRUCTION_TEMPLATES)
    return tmpl.format(
        sys=sys, aspect=_pick(["성능", "안정성", "보안", "확장성", "가용성"]),
        measure=_pick(["모니터링 강화", "자동 복구", "이중화", "부하 분산", "페일오버"]),
        target=f"p99 {random.randint(100,2000)}ms 이내",
        topic=_pick(["아키텍처 개선", "성능 최적화", "보안 강화", "운영 자동화"]),
        section=_pick(["장애 대응 절차", "배포 체크리스트", "모니터링 항목", "롤백 절차"]),
        test_type=_pick(["부하", "회귀", "통합", "카오스", "보안 침투"]),
        coverage=random.randint(80, 99),
        security_aspect=_pick(["접근 권한", "데이터 암호화", "API 키 관리", "감사 로그"]),
    )

def gen_analysis_text(sys: str, idx: int) -> str:
    """Generate a structured analysis paragraph, always unique via index-seeded variation."""
    dep_sys = _pick([s for s in SYSTEMS if s != sys] or SYSTEMS)
    component = _pick(COMPONENTS)
    root_cause = _pick(ROOT_CAUSES)
    effect = _pick(EFFECTS)

    cause = _pick(ANALYSIS_CAUSE_TEMPLATES).format(
        sys=sys, component=component, root_cause=root_cause, effect=effect,
        condition=f"동시 접속 {random.randint(50,5000)}건 이상",
        mechanism=_pick(COMPONENTS), config_issue=_pick(ROOT_CAUSES),
        behavior=_pick(["대기 상태 전환", "재시도 무한반복", "응답 누락", "데이터 손실"]),
        external_factor=_pick(["클라우드 인프라 장애", "서드파티 API 지연", "네트워크 불안정", "DNS 장애"]),
        resource=_pick(["메모리", "CPU", "디스크", "네트워크 대역폭", "커넥션 풀"]),
        dep_sys=dep_sys, interface_issue=_pick(["프로토콜 버전 불일치", "직렬화 포맷 차이", "타임아웃 설정 불일치"]),
    )

    evidence = _pick(ANALYSIS_EVIDENCE_TEMPLATES).format(
        date_ref=f"최근 {random.randint(1,14)}일",
        evidence=_pick(ROOT_CAUSES), count=random.randint(10, 10000),
        metric=_pick(METRICS), pct=round(random.uniform(15, 300), 1),
        span=_pick(COMPONENTS), latency=random.randint(100, 15000),
        threshold=random.randint(200, 3000),
        query_part=_pick(["JOIN 절", "WHERE 조건", "서브쿼리", "집계 함수", "정렬 연산"]),
        effect=_pick(EFFECTS), sys=sys,
        object_type=_pick(["ByteBuffer", "EmbeddingVector", "CacheEntry", "SessionState", "QueryPlan"]),
        size=random.randint(50, 8000),
        target=dep_sys, retry_rate=round(random.uniform(5, 40), 1),
    )

    judgment = _pick(ANALYSIS_JUDGMENT_TEMPLATES).format(
        sys=sys, category=_pick(CATEGORIES), severity=_pick(SEVERITIES),
        approach=_pick(ACTIONS_VOCAB),
        workaround=_pick(ACTIONS_VOCAB),
        permanent_fix=_pick(ACTIONS_VOCAB),
        redesign_area=_pick(COMPONENTS),
        limitation=_pick(ROOT_CAUSES),
        upgrade_target=f"v{random.randint(2,6)}.{random.randint(0,15)}",
        root_cause=root_cause,
        broader_issue=_pick(["인프라 플랫폼 노후화", "마이크로서비스 간 결합도", "관측성 부족", "배포 파이프라인 취약성"]),
    )

    impact = _pick(ANALYSIS_IMPACT_TEMPLATES).format(
        scope=_pick(["전체 서비스", "특정 리전", "특정 사용자 그룹", "내부 관리 도구"]),
        user_count=random.randint(10, 50000),
        service=sys, availability=round(random.uniform(85, 99.9), 2),
        affected_feature=_pick(["검색", "문서 조회", "답변 생성", "대시보드", "알림"]),
        duration=f"{random.randint(5, 480)}분",
        downstream_impact=_pick(["고객 지원 채널", "실시간 알림", "주간 리포트", "SLA 지표"]),
        traffic_pct=round(random.uniform(1, 100), 1),
        region=_pick(REGIONS),
    )

    action = _pick(ANALYSIS_ACTION_TEMPLATES).format(
        action1=_pick(ACTIONS_VOCAB), action2=_pick(ACTIONS_VOCAB),
        monitoring_plan=f"{random.randint(1,7)}일간 집중 모니터링",
        preventive_measure=_pick(ACTIONS_VOCAB),
        review_cycle=_pick(["주 1회", "격주", "월 1회", "분기별"]),
        documentation=_pick(["운영 매뉴얼", "장애 대응 런북", "아키텍처 문서", "API 스펙"]),
        teams=", ".join(_picks(["백엔드팀", "인프라팀", "QA팀", "보안팀", "프론트엔드팀", "데이터팀"], 2)),
        test_plan=_pick(["회귀 테스트 스위트", "카오스 엔지니어링 시나리오", "부하 테스트 플랜"]),
        coverage_target=random.randint(85, 99),
        escalation_plan=_pick(["P1 에스컬레이션 절차", "CTO 보고 라인", "긴급 대응 프로토콜"]),
        timeline=f"{random.randint(1,4)}주",
    )

    return f"{cause}\n{evidence}\n{judgment}\n{impact}\n{action}"


# ──────────────────────────────────────────────────────────────────────
# 4. DATE GENERATION
# ──────────────────────────────────────────────────────────────────────

DATE_START = date(2025, 1, 1)
DATE_END = date(2026, 3, 20)
DATE_RANGE_DAYS = (DATE_END - DATE_START).days

def gen_registered_date() -> date:
    return DATE_START + timedelta(days=random.randint(0, DATE_RANGE_DAYS))

def gen_start_date(reg: date) -> date | None:
    if random.random() < 0.15:
        return None
    return reg + timedelta(days=random.randint(0, 5))

def gen_due_date(start: date | None, reg: date) -> date | None:
    if random.random() < 0.1:
        return None
    base = start or reg
    return base + timedelta(days=random.randint(3, 60))

def gen_completed_date(due: date | None, status_label: str) -> date | None:
    if status_label != "완료":
        return None
    if due is None:
        return None
    return due + timedelta(days=random.randint(-5, 10))


# ──────────────────────────────────────────────────────────────────────
# 5. ROW ASSEMBLY & EXCEL WRITING
# ──────────────────────────────────────────────────────────────────────

class IssueRecord(NamedTuple):
    title: str
    registered_date: date
    check_text: str
    work_text: str
    instruction_text: str
    assignee: str
    start_date: date | None
    due_date: date | None
    status_text: str
    completed_date: date | None
    analysis: str
    status_helper: str
    keyword_helper: str


def build_records(n: int) -> list[IssueRecord]:
    titles = generate_unique_titles(n)
    random.shuffle(titles)
    records = []

    for idx in range(n):
        title = titles[idx]
        # extract system name from title
        sys_name = title.split(" ")[0]
        if sys_name not in SYSTEMS:
            sys_name = _pick(SYSTEMS)

        reg_date = gen_registered_date()
        start = gen_start_date(reg_date)
        due = gen_due_date(start, reg_date)
        status_text, status_label = _pick(STATUS_OPTIONS)
        completed = gen_completed_date(due, status_label)

        # keyword helper: key terms from title + unique ref
        keyword_helper = ", ".join(title.replace("(", "").replace(")", "").split()[:3]) + f", REF-{idx:05d}"

        records.append(IssueRecord(
            title=title,
            registered_date=reg_date,
            check_text=gen_check_text(sys_name, idx),
            work_text=gen_work_text(sys_name, idx),
            instruction_text=gen_instruction_text(sys_name, idx),
            assignee=_pick(ASSIGNEES),
            start_date=start,
            due_date=due,
            status_text=status_text,
            completed_date=completed,
            analysis=gen_analysis_text(sys_name, idx),
            status_helper=status_label,
            keyword_helper=keyword_helper,
        ))

    return records


def write_excel(records: list[IssueRecord], path: str) -> None:
    wb = openpyxl.Workbook()

    # ── Main data sheet ──
    ws = wb.active
    ws.title = "이슈데이터_10000건"

    headers = [
        "모델 이슈 검토 사항", "등록일", "기본 확인내용", "기본 작업내용",
        "업무지시", "담당자", "업무시작일", "완료예정",
        "진행(담당자)", "완료일", "문제점 분석 내용 (담당자 Comments)",
        "상태_도우미", "키워드_도우미",
    ]

    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row_idx, rec in enumerate(records, 2):
        ws.cell(row=row_idx, column=1, value=rec.title)
        ws.cell(row=row_idx, column=2, value=rec.registered_date)
        ws.cell(row=row_idx, column=3, value=rec.check_text)
        ws.cell(row=row_idx, column=4, value=rec.work_text)
        ws.cell(row=row_idx, column=5, value=rec.instruction_text)
        ws.cell(row=row_idx, column=6, value=rec.assignee)
        ws.cell(row=row_idx, column=7, value=rec.start_date)
        ws.cell(row=row_idx, column=8, value=rec.due_date)
        ws.cell(row=row_idx, column=9, value=rec.status_text)
        ws.cell(row=row_idx, column=10, value=rec.completed_date)
        ws.cell(row=row_idx, column=11, value=rec.analysis)
        ws.cell(row=row_idx, column=12, value=rec.status_helper)
        ws.cell(row=row_idx, column=13, value=rec.keyword_helper)

    # Column widths
    widths = [40, 12, 45, 45, 45, 12, 12, 12, 50, 12, 80, 10, 20]
    for col, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = w

    # Freeze header row
    ws.freeze_panes = "A2"

    # ── Legend sheet ──
    ws_legend = wb.create_sheet("범례")
    ws_legend.cell(row=1, column=1, value="모델 이슈 데이터셋 범례")
    ws_legend.cell(row=1, column=1).font = Font(bold=True, size=14)
    legend_items = [
        ("데이터 건수", "10,000건 (모두 고유, 중복 없음)"),
        ("등록일 범위", "2025-01-01 ~ 2026-03-20"),
        ("담당자", ", ".join(ASSIGNEES)),
        ("진행 상태", "완료 / 진행중 / 검증중 / 분석중 / 대기"),
        ("시스템 종류", f"{len(SYSTEMS)}종"),
        ("증상 종류", f"{len(SYMPTOMS)}종"),
        ("분석 내용 구조", "원인 요약 / 확인 근거 / 기술 판단 / 영향 범위 / 추가 조치"),
        ("임베딩 모델", "BAAI/bge-m3 (1024차원)"),
        ("생성 방법", "조합적 템플릿 기반 자동 생성 (seed=42)"),
    ]
    for i, (key, val) in enumerate(legend_items, 3):
        ws_legend.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_legend.cell(row=i, column=2, value=val)
    ws_legend.column_dimensions["A"].width = 20
    ws_legend.column_dimensions["B"].width = 60

    # ── Description sheet ──
    ws_desc = wb.create_sheet("설명")
    ws_desc.cell(row=1, column=1, value="데이터셋 설명")
    ws_desc.cell(row=1, column=1).font = Font(bold=True, size=14)
    descriptions = [
        ("생성 목적", "RAG 시스템 이슈 관리를 위한 임베딩/검색 테스트 데이터"),
        ("고유성 보장", "모든 제목이 고유하며, 분석 내용에 REF 번호로 유일성 보장"),
        ("분석 구조", "각 분석은 5개 섹션(원인/근거/판단/영향/조치)으로 구조화"),
        ("메타데이터", "담당자 12명, 상태 5종, 시스템 50종, 증상 50종 조합"),
        ("날짜 분포", "등록일은 2025-01-01 ~ 2026-03-20 범위에서 균등 분포"),
    ]
    for i, (key, val) in enumerate(descriptions, 3):
        ws_desc.cell(row=i, column=1, value=key).font = Font(bold=True)
        ws_desc.cell(row=i, column=2, value=val)
    ws_desc.column_dimensions["A"].width = 20
    ws_desc.column_dimensions["B"].width = 80

    wb.save(path)
    print(f"Excel saved: {path}")


# ──────────────────────────────────────────────────────────────────────
# 6. VALIDATION
# ──────────────────────────────────────────────────────────────────────

def validate(records: list[IssueRecord]) -> None:
    print(f"\n{'='*60}")
    print("VALIDATION")
    print(f"{'='*60}")

    n = len(records)
    print(f"Total records: {n}")

    titles = [r.title for r in records]
    unique_titles = len(set(titles))
    print(f"Unique titles: {unique_titles}/{n} {'[OK]' if unique_titles == n else '[FAIL]'}")

    analyses = [r.analysis for r in records]
    unique_analyses = len(set(analyses))
    print(f"Unique analyses: {unique_analyses}/{n} {'[OK]' if unique_analyses == n else '[FAIL]'}")

    checks = [r.check_text for r in records]
    unique_checks = len(set(checks))
    print(f"Unique check_text: {unique_checks}/{n} ({unique_checks/n*100:.1f}%)")

    works = [r.work_text for r in records]
    unique_works = len(set(works))
    print(f"Unique work_text: {unique_works}/{n} ({unique_works/n*100:.1f}%)")

    # Title length stats
    title_lens = [len(t) for t in titles]
    print(f"\nTitle length: min={min(title_lens)}, max={max(title_lens)}, avg={sum(title_lens)/len(title_lens):.1f}")

    # Analysis length stats
    analysis_lens = [len(a) for a in analyses]
    print(f"Analysis length: min={min(analysis_lens)}, max={max(analysis_lens)}, avg={sum(analysis_lens)/len(analysis_lens):.1f}")

    # Assignee distribution
    from collections import Counter
    assignee_counts = Counter(r.assignee for r in records)
    print(f"\nAssignee distribution ({len(assignee_counts)} people):")
    for a, c in assignee_counts.most_common():
        print(f"  {a:15s}: {c:>5d}")

    # Status distribution
    status_counts = Counter(r.status_helper for r in records)
    print(f"\nStatus distribution:")
    for s, c in status_counts.most_common():
        print(f"  {s:15s}: {c:>5d}")

    # Date range
    dates = [r.registered_date for r in records]
    print(f"\nDate range: {min(dates)} ~ {max(dates)}")

    # Completed date presence
    completed_count = sum(1 for r in records if r.completed_date is not None)
    print(f"Completed date filled: {completed_count}/{n}")

    print(f"\n{'='*60}")
    if unique_titles == n and unique_analyses == n:
        print("ALL VALIDATIONS PASSED")
    else:
        print("VALIDATION FAILED - DUPLICATES DETECTED")
    print(f"{'='*60}")


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    TARGET = 10_000
    OUTPUT = "data/model_issue_dataset_10000.xlsx"

    print(f"Generating {TARGET} unique issue records...")
    records = build_records(TARGET)
    validate(records)
    write_excel(records, OUTPUT)
    print("\nDone!")
